import math

import plotly.graph_objects as go
from conftest import load_scenario
from openpyxl import load_workbook

import clashless as cl


def _solve_parallel_rooms_unbounded():
    scenario = load_scenario("parallel_rooms_unbounded")
    schedule = cl.Schedule(
        scenario.participants,
        scenario.sessions,
        scenario.unavailability,
        1,
        1,
    ).solve()
    return scenario, schedule


def test_plot_schedule_returns_a_figure_with_one_cell_per_presentation():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule, scenario.participants, scenario.sessions, n_days=1, n_slots=1
    )

    assert isinstance(fig, go.Figure)
    heatmap = fig.data[0]

    occupied = sum(1 for row in heatmap.z for value in row if not math.isnan(value))
    assert occupied == len(schedule)

    # all three (distinct-session) presentations should get their own column
    assert set(heatmap.x) == set(scenario.sessions.data.values())


def test_plot_schedule_labels_each_cell_with_its_presentation_id():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule, scenario.participants, scenario.sessions, n_days=1, n_slots=1
    )
    heatmap = fig.data[0]

    labelled_ids = {value for row in heatmap.text for value in row if value}
    assert labelled_ids == set(schedule)


def test_plot_schedule_hover_includes_id_and_participants():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule, scenario.participants, scenario.sessions, n_days=1, n_slots=1
    )
    heatmap = fig.data[0]

    presentation_id = next(iter(schedule))
    participants = scenario.participants.data[presentation_id]
    hover_entries = [value for r in heatmap.customdata for value in r if value]
    matching = [text for text in hover_entries if f"ID: {presentation_id}" in text]

    assert len(matching) == 1
    for participant in participants:
        assert participant in matching[0]


def test_plot_schedule_uses_given_slot_labels():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        slot_labels={1: "09:00"},
    )

    assert list(fig.data[0].y) == ["Day 1 · 09:00"]


def test_plot_schedule_uses_given_day_labels():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        day_labels={1: "Mon"},
    )

    assert list(fig.data[0].y) == ["Mon · Slot 1"]


def test_plot_schedule_highlights_presentations_with_matching_participants():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        highlight_participants=["Alice Smith"],
    )
    heatmap = fig.data[0]

    highlighted_id = next(
        presentation_id
        for presentation_id, people in scenario.participants.data.items()
        if "Alice Smith" in people
    )

    for i, row in enumerate(heatmap.z):
        for j, value in enumerate(row):
            if math.isnan(value):
                continue
            is_highlighted_cell = heatmap.text[i][j] == highlighted_id
            expected_z = (
                cl.plotting.HIGHLIGHT_Z if is_highlighted_cell else cl.plotting.NORMAL_Z
            )
            assert value == expected_z

    # id labels and hover text stay populated regardless of highlight state
    assert {value for row in heatmap.text for value in row if value} == set(schedule)


def test_plot_schedule_without_highlight_participants_uses_normal_color_for_all_cells():
    scenario, schedule = _solve_parallel_rooms_unbounded()

    fig = cl.plot_schedule(
        schedule, scenario.participants, scenario.sessions, n_days=1, n_slots=1
    )
    heatmap = fig.data[0]

    occupied_values = [
        value for row in heatmap.z for value in row if not math.isnan(value)
    ]
    assert occupied_values == [cl.plotting.NORMAL_Z] * len(occupied_values)


def test_export_schedule_to_excel_writes_full_details_per_cell(tmp_path):
    scenario, schedule = _solve_parallel_rooms_unbounded()
    path = tmp_path / "schedule.xlsx"

    cl.export_schedule_to_excel(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        path=path,
    )

    workbook = load_workbook(path)
    sheet = workbook.active

    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert set(header_row[2:]) == set(scenario.sessions.data.values())

    filled_cells = [
        cell.value
        for row in sheet.iter_rows(min_row=2)
        for cell in row
        if cell.column > 2 and cell.value
    ]
    assert len(filled_cells) == len(schedule)

    presentation_id = next(iter(schedule))
    participants = scenario.participants.data[presentation_id]
    (matching,) = [text for text in filled_cells if f"id: {presentation_id}" in text]
    for participant in participants:
        assert participant in matching


def test_export_schedule_to_excel_cell_text_uses_lowercase_id_and_no_day_slot_line(
    tmp_path,
):
    scenario, schedule = _solve_parallel_rooms_unbounded()
    path = tmp_path / "schedule.xlsx"

    cl.export_schedule_to_excel(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        path=path,
    )

    workbook = load_workbook(path)
    sheet = workbook.active

    presentation_id = next(iter(schedule))
    filled_cells = [
        cell.value
        for row in sheet.iter_rows(min_row=2)
        for cell in row
        if cell.column > 2 and cell.value
    ]
    (matching,) = [text for text in filled_cells if f"id: {presentation_id}" in text]

    assert matching.startswith(f"id: {presentation_id}")
    assert "Day" not in matching
    assert "slot" not in matching
    assert matching.count("\n") == 2  # id, Participants, Session - three lines


def test_export_schedule_to_excel_has_separate_day_and_slot_header_columns(tmp_path):
    scenario, schedule = _solve_parallel_rooms_unbounded()
    path = tmp_path / "schedule.xlsx"

    cl.export_schedule_to_excel(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        path=path,
    )

    workbook = load_workbook(path)
    sheet = workbook.active

    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row[0] == "Day"
    assert header_row[1] == "Slot"
    assert set(header_row[2:]) == set(scenario.sessions.data.values())

    data_row = next(sheet.iter_rows(min_row=2, max_row=2))
    assert data_row[0].value == "Day 1"
    assert data_row[1].value == "Slot 1"


def test_export_schedule_to_excel_uses_given_day_and_slot_labels(tmp_path):
    scenario, schedule = _solve_parallel_rooms_unbounded()
    path = tmp_path / "schedule.xlsx"

    cl.export_schedule_to_excel(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        path=path,
        day_labels={1: "Mon"},
        slot_labels={1: "09:00"},
    )

    workbook = load_workbook(path)
    sheet = workbook.active
    data_row = next(sheet.iter_rows(min_row=2, max_row=2))

    assert data_row[0].value == "Mon"
    assert data_row[1].value == "09:00"


def test_export_schedule_to_excel_has_no_fill_styling(tmp_path):
    scenario, schedule = _solve_parallel_rooms_unbounded()
    path = tmp_path / "schedule.xlsx"

    cl.export_schedule_to_excel(
        schedule,
        scenario.participants,
        scenario.sessions,
        n_days=1,
        n_slots=1,
        path=path,
    )

    workbook = load_workbook(path)
    sheet = workbook.active

    for row in sheet.iter_rows():
        for cell in row:
            assert cell.fill.fill_type is None
