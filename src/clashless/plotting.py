from collections import Counter

import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Validated dataviz palette (see project's dataviz skill / palette.md). Normally this
# is a single-series chart (occupied vs. empty slot), needing only the accent hue and
# chart chrome tokens - no categorical CVD comparison applies. highlight_participants
# introduces a second, genuinely categorical state (highlighted vs. not), so
# HIGHLIGHT_COLOR is deliberately the palette's categorical slot 2 (orange) -
# ACCENT_COLOR is already slot 1 (blue), and slots 1<->2 are that palette's
# best-separated adjacent pair.
ACCENT_COLOR = "#2a78d6"
HIGHLIGHT_COLOR = "#eb6834"  # categorical slot 2 (orange)
SURFACE_COLOR = "#fcfcfb"
PRIMARY_TEXT = "#0b0b0b"
MUTED_TEXT = "#898781"
AXIS_COLOR = "#c3c2b7"

FONT_FAMILY = 'system-ui, -apple-system, "Segoe UI", sans-serif'

# Heatmap z-values are discrete markers, never interpolated between - each cell is
# exactly one of these two states, so the two-stop colorscale below always resolves
# to an exact endpoint color.
NORMAL_Z = 1
HIGHLIGHT_Z = 2


def _slot_label(slot, slot_labels):
    return (slot_labels or {}).get(slot, f"Slot {slot}")


def _day_label(day, day_labels):
    return (day_labels or {}).get(day, f"Day {day}")


def _build_grid(schedule, participants, sessions, n_days, n_slots):
    """Lay a solved schedule out on a room x time grid.

    Shared by plot_schedule and export_schedule_to_excel so both render
    exactly the same layout. Returns the ordered session (room) labels, the
    ordered chronological (day, slot) row keys, and a row_keys x
    session_order matrix of per-presentation detail dicts (None where
    nothing is scheduled). Callers format row_keys into display labels
    themselves, since plot_schedule needs one combined string per row while
    export_schedule_to_excel needs day and slot as separate columns.
    """
    session_order = [
        session
        for session, _ in Counter(
            sessions.data[presentation_id] for presentation_id in schedule
        ).most_common()
    ]

    row_keys = [
        (day, slot) for day in range(1, n_days + 1) for slot in range(1, n_slots + 1)
    ]
    row_index = {key: i for i, key in enumerate(row_keys)}
    column_index = {session: j for j, session in enumerate(session_order)}

    cells = [[None] * len(session_order) for _ in row_keys]
    for presentation_id, entry in schedule.items():
        day, slot = entry["day"], entry["slot"]
        session = sessions.data[presentation_id]
        i = row_index[day, slot]
        j = column_index[session]
        cells[i][j] = {
            "id": presentation_id,
            "participants": participants.data[presentation_id],
            "session": session,
            "day": day,
            "slot": slot,
        }

    return session_order, row_keys, cells


def plot_schedule(  # noqa: PLR0917
    schedule,
    participants,
    sessions,
    n_days,
    n_slots,
    slot_labels=None,
    day_labels=None,
    highlight_participants=None,
):
    """Render an interactive timetable for a solved Schedule.

    Sessions (rooms/tracks) run along the x-axis and chronological (day, slot)
    rows run down the y-axis - both are positional, so they carry room/time
    identity without needing a color per session. Each filled cell is one
    scheduled presentation in the single accent color, labelled with its
    presentation id; hovering it shows the full detail - id, participants,
    and session. Empty cells are real gaps: a session only owns one room for
    the whole day it runs, so a free slot in its column means nothing is
    scheduled there, not missing data. `slot_labels`/`day_labels`, if given,
    are optional `{slot_number: "09:00"}`/`{day_number: "Mon"}` maps for
    cosmetic row labels - purely for display, entirely decoupled from
    scheduling. `highlight_participants`, if given, is a list of one or more
    participant names; any presentation whose participants include one or
    more of them is rendered in a single distinct highlight color instead of
    the normal accent color (not a color per name - just highlighted vs.
    not).
    """
    session_order, row_keys, cells = _build_grid(
        schedule, participants, sessions, n_days, n_slots
    )
    row_labels = [
        f"{_day_label(day, day_labels)} · {_slot_label(slot, slot_labels)}"
        for day, slot in row_keys
    ]
    highlight_set = set(highlight_participants or ())

    z = [[float("nan")] * len(session_order) for _ in row_labels]
    id_labels = [[""] * len(session_order) for _ in row_labels]
    hover_text = [[""] * len(session_order) for _ in row_labels]

    for i, row_cells in enumerate(cells):
        for j, info in enumerate(row_cells):
            if info is None:
                continue
            is_highlighted = bool(highlight_set & set(info["participants"]))
            z[i][j] = HIGHLIGHT_Z if is_highlighted else NORMAL_Z
            id_labels[i][j] = str(info["id"])
            hover_text[i][j] = (
                f"<b>ID: {info['id']}</b><br>"
                f"Participants: {', '.join(str(p) for p in info['participants'])}<br>"
                f"Session: {info['session']}<br>"
                f"Day {info['day']}, slot {info['slot']}"
            )

    fig = go.Figure(
        go.Heatmap(
            z=z,
            x=session_order,
            y=row_labels,
            text=id_labels,
            texttemplate="%{text}",
            textfont=dict(color="#ffffff", size=11, family=FONT_FAMILY),
            customdata=hover_text,
            hovertemplate="%{customdata}<extra></extra>",
            hoverongaps=False,
            colorscale=[[0, ACCENT_COLOR], [1, HIGHLIGHT_COLOR]],
            zmin=NORMAL_Z,
            zmax=HIGHLIGHT_Z,
            showscale=False,
            xgap=3,
            ygap=3,
        )
    )

    fig.update_layout(
        title=dict(
            text=(
                "Conference schedule"
                f"<br><sup style='color:{MUTED_TEXT}'>"
                "Each cell is a scheduled presentation (labelled by id) — hover for "
                "full details"
                "</sup>"
            ),
            x=0.02,
            xanchor="left",
        ),
        xaxis=dict(
            title="Session (room)",
            side="top",
            showgrid=False,
            linecolor=AXIS_COLOR,
            tickfont=dict(color=MUTED_TEXT),
        ),
        yaxis=dict(
            title="Day · slot",
            autorange="reversed",
            showgrid=False,
            linecolor=AXIS_COLOR,
            tickfont=dict(color=MUTED_TEXT),
        ),
        plot_bgcolor=SURFACE_COLOR,
        paper_bgcolor=SURFACE_COLOR,
        font=dict(family=FONT_FAMILY, color=PRIMARY_TEXT),
        margin=dict(l=160, r=40, t=100, b=20),
        height=max(400, 26 * len(row_labels) + 160),
        width=max(520, 140 * len(session_order) + 200),
    )

    return fig


DATA_COLUMN_WIDTH = 36  # openpyxl width units ~= characters of the default font
DATA_FONT_SIZE = 10
POINTS_PER_LINE = 15  # ~ one wrapped line at DATA_FONT_SIZE, including leading
ROW_PADDING_POINTS = 8


def _cell_text(info):
    participants_text = ", ".join(str(p) for p in info["participants"])
    return (
        f"id: {info['id']}\n"
        f"Participants: {participants_text}\n"
        f"Session: {info['session']}"
    )


def _wrapped_line_count(text, column_width):
    # Estimates how many visual lines `text` wraps to at `column_width` characters
    # per line - deliberately a slight over-estimate (never under), since a row
    # that's a touch taller than it needs to be is far less broken than one that
    # clips or overlaps the row below it.
    chars_per_line = max(1, int(column_width))
    return sum(-(-len(line) // chars_per_line) for line in text.split("\n"))


DAY_SLOT_COLUMN_WIDTH = 14


def export_schedule_to_excel(  # noqa: PLR0917
    schedule,
    participants,
    sessions,
    n_days,
    n_slots,
    path,
    slot_labels=None,
    day_labels=None,
):
    """Write a solved schedule to an .xlsx file, mirroring plot_schedule's layout.

    A spreadsheet has no hover, so every occupied cell's full detail - id,
    participants, and session - is written directly as wrapped text, not
    just a compact label; day and slot are shown in their own leading
    columns instead of repeated inside every cell. Row heights are sized per
    row to fit whichever of that row's cells wraps to the most lines, so
    long names never get clipped or bleed into the row below. There is no
    cell coloring - identity comes entirely from the columns and text.
    `slot_labels`/`day_labels`, if given, are optional cosmetic maps, same
    as in plot_schedule.
    """
    session_order, row_keys, cells = _build_grid(
        schedule, participants, sessions, n_days, n_slots
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"

    header_font = Font(bold=True)
    row_header_font = Font(bold=True, size=DATA_FONT_SIZE)
    data_font = Font(size=DATA_FONT_SIZE)
    wrap_top = Alignment(wrap_text=True, vertical="top")

    sheet.cell(row=1, column=1, value="Day").font = header_font
    sheet.cell(row=1, column=2, value="Slot").font = header_font

    for j, session in enumerate(session_order, start=3):
        header_cell = sheet.cell(row=1, column=j, value=session)
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, (day, slot) in enumerate(row_keys, start=2):
        day_cell = sheet.cell(row=i, column=1, value=_day_label(day, day_labels))
        day_cell.font = row_header_font
        slot_cell = sheet.cell(row=i, column=2, value=_slot_label(slot, slot_labels))
        slot_cell.font = row_header_font

        max_lines = 1
        for j, info in enumerate(cells[i - 2], start=3):
            if info is None:
                continue
            text = _cell_text(info)
            excel_cell = sheet.cell(row=i, column=j, value=text)
            excel_cell.font = data_font
            excel_cell.alignment = wrap_top
            max_lines = max(max_lines, _wrapped_line_count(text, DATA_COLUMN_WIDTH))

        row_height = max_lines * POINTS_PER_LINE + ROW_PADDING_POINTS
        sheet.row_dimensions[i].height = row_height

    sheet.column_dimensions[get_column_letter(1)].width = DAY_SLOT_COLUMN_WIDTH
    sheet.column_dimensions[get_column_letter(2)].width = DAY_SLOT_COLUMN_WIDTH
    for j in range(3, len(session_order) + 3):
        sheet.column_dimensions[get_column_letter(j)].width = DATA_COLUMN_WIDTH

    sheet.freeze_panes = "C2"

    workbook.save(path)
